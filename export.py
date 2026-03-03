"""
Main export script: reads Belfius CSV from input/, cleans, categorizes,
and outputs Excel dashboard + clean CSV.
"""

import json
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname, get_column_letter
from openpyxl.styles import Font

INPUT_DIR = Path("input")
OUTPUT_DIR = Path("output")
CONFIG_FILE = Path("config") / "categories.json"
EXCEL_OUTPUT = OUTPUT_DIR / "clean_dashboard.xlsx"
CSV_OUTPUT = OUTPUT_DIR / "transactions_clean.csv"


# --- Belfius CSV parsing ---

def find_first_csv() -> Path | None:
    """Find first CSV file in input/ directory."""
    if not INPUT_DIR.exists():
        return None
    for f in sorted(INPUT_DIR.iterdir()):
        if f.suffix.lower() == ".csv":
            return f
    return None


def parse_belfius_csv(path: Path) -> pd.DataFrame:
    """Parse Belfius CSV with sep=;, header detection, cp1252/latin1 encoding."""
    for enc in ("cp1252", "latin1", "utf-8"):
        try:
            with open(path, "r", encoding=enc) as f:
                lines = f.readlines()
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError(f"Cannot decode {path}")

    # Find header row (starts with "Compte;Date de comptabilisation")
    header_idx = None
    for i, line in enumerate(lines):
        if line.strip().startswith("Compte;Date de comptabilisation"):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Belfius header row not found")

    df = pd.read_csv(path, sep=";", encoding=enc, skiprows=header_idx, dtype=str)
    return df


def parse_european_amount(s: str) -> float:
    """Parse European number format (e.g. 1.234,56 or -750)."""
    if pd.isna(s) or str(s).strip() == "":
        return 0.0
    s = str(s).strip().replace(" ", "")
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_date(s: str) -> str | None:
    """Parse DD/MM/YYYY to YYYY-MM-DD."""
    if pd.isna(s) or str(s).strip() == "":
        return None
    try:
        dt = datetime.strptime(str(s).strip()[:10], "%d/%m/%Y")
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def clean_and_normalize(df: pd.DataFrame) -> pd.DataFrame:
    """Clean and normalize Belfius data to standard columns."""
    col_map = {
        "Compte": "account_iban",
        "Date de comptabilisation": "booking_date",
        "Numéro d'extrait": "extract_nr",
        "Numéro de transaction": "transaction_nr",
        "Compte contrepartie": "counterparty_account",
        "Nom contrepartie contient": "counterparty",
        "Rue et numéro": "street",
        "Code postal et localité": "city",
        "Transaction": "description",
        "Date valeur": "value_date",
        "Montant": "amount",
        "Devise": "currency",
        "BIC": "bic",
        "Code pays": "country_code",
        "Communications": "communications",
    }
    out = pd.DataFrame()
    for old, new in col_map.items():
        if old in df.columns:
            out[new] = df[old]
        else:
            out[new] = ""

    # Normalize
    out["account_iban"] = out["account_iban"].fillna("").astype(str).str.strip()
    out["booking_date"] = out["booking_date"].apply(parse_date)
    out["value_date"] = out["value_date"].apply(parse_date)
    out["amount"] = out["amount"].apply(parse_european_amount)
    out["currency"] = out["currency"].fillna("EUR").astype(str).str.strip()
    out["counterparty"] = out["counterparty"].fillna("").astype(str).str.strip()
    out["counterparty_account"] = out["counterparty_account"].fillna("").astype(str).str.strip()
    out["description"] = out["description"].fillna("").astype(str).str.strip()
    out["communications"] = out["communications"].fillna("").astype(str).str.strip()

    # raw_type: first part of description (e.g. VIREMENT, PAIEMENT DEBITMASTERCARD)
    out["raw_type"] = out["description"].str.extract(r"^([A-Z\s\*]+)", expand=False).fillna("")
    out["raw_type"] = out["raw_type"].str.strip()

    # direction
    out["direction"] = out["amount"].apply(lambda x: "in" if x >= 0 else "out")

    # month (YYYY-MM)
    out["month"] = out["booking_date"].str[:7] if out["booking_date"].notna().any() else ""

    # quarter (e.g. 2026-Q1)
    def to_quarter(m):
        if pd.isna(m) or str(m).strip() == "" or len(str(m)) < 7:
            return ""
        try:
            y, mo = int(str(m)[:4]), int(str(m)[5:7])
            return f"{y}-Q{(mo - 1) // 3 + 1}"
        except (ValueError, IndexError):
            return ""

    out["quarter"] = out["month"].apply(to_quarter)

    return out


# --- Rule engine ---

def _text(desc: str, counterparty: str) -> str:
    """Combined searchable text."""
    return f" {str(desc).upper()} {str(counterparty).upper()} "


def apply_rules(row: pd.Series, categories: dict) -> tuple[str, str]:
    """Apply smart rules to assign category and subcategory."""
    desc = str(row.get("description", ""))
    cp = str(row.get("counterparty", ""))
    raw = str(row.get("raw_type", ""))
    txt = _text(desc, cp)

    # Order matters: specific rules first. Admin/Capital before broad Marketing to avoid over-matching (e.g. ROOFWANDER in notary/capital refs).
    rules = [
        # Administration - Banking Fees
        (("FRAIS D'EXPEDITION", "FRAIS D EXPEDITION", "TARIFICATION", "AVIS BANCAIRES", "OPERATIONS DE DEBIT"), "Administration", "Banking Fees"),
        # Administration - Accounting & Fiduciary
        (("FIDUCIAIRE", "UCM TECHNICS"), "Administration", "Accounting & Fiduciary"),
        # Administration - Legal & Compliance (before Partnerships – NOTAIRE can appear with ROOFWANDER in ref)
        (("NOTAIRE", "ETUDE DU NOTAIRE"), "Administration", "Legal & Compliance"),
        # Administration - Admin Subscriptions
        (("LES MANALAS", "ABONNEMENT"), "Administration", "Admin Subscriptions"),
        # Administration - Office & Supplies
        (("AMAZON", "BIZAY", "ACHAT CARBURANT", "EUSKADI LOW COST", "CARBURANT", "BANCONTACT", "CRF EXP"), "Administration", "Office & Supplies"),
        # Capital & Financing – before Marketing (INVESTISSEMENT/VERSAMENTO in capital refs)
        (("INVESTISSEMENT", "VERSAMENTO CAPITALE", "CAPITALE SOCIALE", "ROOFWANDER ACCOUNT"), "Capital & Financing", "Founder Contributions"),
        # Marketing - Paid Advertising
        (("GOOGLE ADS", "GOOGLE*ADS", "MICROSOFT*ADS", "MICROSOFT*ADVERTISING", "MICROSOFT IRELAND", "LEBONCOIN"), "Marketing", "Paid Advertising"),
        # Marketing - Events & Offline Marketing
        (("MAGNIS GROUP", "CMONEVENT", "EMPLACEMENT EXPOSANT", "SARL CMONEVENT", "MECHILLI", "SOCIAL MEDIA", "DACHZELTNOMADEN", "IT'S SO GOOD", "YOGI EDITIONS", "PERSPECTIVES VALLEY"), "Marketing", "Events & Offline Marketing"),
        # Technology - Hosting & Infrastructure
        (("RENDER.COM", "RENDER."), "Technology", "Hosting & Infrastructure"),
        # Technology - Marketplace Platform
        (("SHARETRIBE",), "Technology", "Marketplace Platform"),
        # Technology - SaaS Tools
        (("PROJECTIONHUB", "CLEVERBRIDGE"), "Technology", "SaaS Tools"),
        # Technology - Payments & Transaction Fees
        (("FRAIS DE TRAITEMENT", "COURS DE CHANGE"), "Technology", "Payments & Transaction Fees"),
        # Revenue - incoming (after Capital so investissement/capitale matches first)
        (("VERSEMENT DE",), "Revenue", "Other Operating Revenue"),
    ]

    for keywords, cat, sub in rules:
        for kw in keywords:
            if kw in txt or kw in raw:
                if cat in categories:
                    subs = categories[cat]
                    if sub in subs:
                        return cat, sub
                    # Known category, unknown subcategory: leave empty for user to choose
                    return cat, ""
                return "", ""

    # No rule matched: leave both empty for user to choose
    return "", ""


def _add_excel_pivot_table(filepath: str, n_tx: int) -> None:
    """Add real Excel pivot table + pivot chart using xlwings (requires Excel installed)."""
    try:
        import xlwings as xw
    except ImportError:
        print("Note: Install xlwings for real pivot tables: pip install xlwings")
        return

    try:
        app = xw.App(visible=False)
        wb = app.books.open(filepath)
        ws_tx = wb.sheets["Transactions"]
        ws_pivot = wb.sheets["Pivot"]

        # Source: Transactions data range (header + data rows)
        last_col = ws_tx.range((1, 1)).end("right").column
        source_range = ws_tx.range((1, 1), (n_tx + 1, last_col))
        dest_cell = ws_pivot.range("A1")

        # Create pivot cache and pivot table
        pc = wb.api.PivotCaches().Create(
            SourceType=xw.constants.PivotTableSourceType.xlDatabase,
            SourceData=source_range.api,
        )
        pt = pc.CreatePivotTable(
            TableDestination=dest_cell.api,
            TableName="PivotTable1",
        )

        # Configure fields: Rows = Quarter, Month | Columns = Category | Values = Sum of amount
        pt.PivotFields("quarter").Orientation = xw.constants.PivotFieldOrientation.xlRowField
        pt.PivotFields("quarter").Position = 1
        pt.PivotFields("month").Orientation = xw.constants.PivotFieldOrientation.xlRowField
        pt.PivotFields("month").Position = 2
        pt.PivotFields("category").Orientation = xw.constants.PivotFieldOrientation.xlColumnField
        pt.AddDataField(
            pt.PivotFields("amount"),
            "Sum of Amount",
            xw.constants.ConsolidationFunction.xlSum,
        )

        # Create pivot chart (dynamic, linked to pivot table)
        pt_range = pt.TableRange2
        chart_obj = ws_pivot.api.ChartObjects().Add(350, 20, 400, 250)
        chart_obj.Chart.SetSourceData(Source=pt_range)
        chart_obj.Chart.ChartType = xw.constants.ChartType.xlColumnClustered
        chart_obj.Chart.HasTitle = True
        chart_obj.Chart.ChartTitle.Text = "Amount by Month & Category"

        wb.save()
        wb.close()
        app.quit()
    except Exception as e:
        print(f"Note: Could not create pivot table (Excel required): {e}")
        try:
            app.quit()
        except NameError:
            pass


def _cat_to_range_name(cat: str) -> str:
    """Convert category name to valid Excel range name (no spaces, &, etc.)."""
    return str(cat).replace(" & ", "_").replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")


# --- Excel dashboard ---

def create_excel_dashboard(df: pd.DataFrame, categories: dict) -> None:
    """Create clean_dashboard.xlsx with Subcategories, Lists, Transactions, Summary, Pivot."""
    from openpyxl.workbook.defined_name import DefinedName

    wb = Workbook()

    # --- Subcategories sheet: one column per category for named ranges ---
    ws_subs = wb.active
    ws_subs.title = "Subcategories"
    for col_idx, (cat, subs) in enumerate(categories.items(), 1):
        ws_subs.cell(1, col_idx, cat)
        for row_idx, sub in enumerate(subs, 2):
            ws_subs.cell(row_idx, col_idx, sub)
        # Create named range for this category (valid name for Excel)
        range_name = _cat_to_range_name(cat)
        end_row = 1 + len(subs)
        col_ltr = get_column_letter(col_idx)
        ref = f"{quote_sheetname(ws_subs.title)}!${col_ltr}$2:${col_ltr}${end_row}"
        wb.defined_names.add(DefinedName(range_name, attr_text=ref))

    # Lists sheet (reference)
    ws_lists = wb.create_sheet("Lists", 1)
    ws_lists["A1"] = "Categories (from config)"
    ws_lists["A1"].font = Font(bold=True)
    for row, (cat, subs) in enumerate(categories.items(), 2):
        ws_lists[f"A{row}"] = cat
        ws_lists[f"B{row}"] = ", ".join(subs) if subs else "-"

    # --- Transactions sheet ---
    ws_tx = wb.create_sheet("Transactions", 2)
    cols = list(df.columns) + ["category", "subcategory"]
    cat_col = cols.index("category") + 1
    sub_col = cols.index("subcategory") + 1
    amount_col = cols.index("amount") + 1
    n_tx = len(df)

    for c, col in enumerate(cols, 1):
        ws_tx.cell(1, c, col)
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            ws_tx.cell(r, c, val)

    # Column letters for formulas
    cat_letter = get_column_letter(cat_col)
    sub_letter = get_column_letter(sub_col)
    amount_letter = get_column_letter(amount_col)
    tx_data_end = 1 + n_tx

    # Category dropdown
    cat_options = ",".join(categories.keys())
    dv_cat = DataValidation(type="list", formula1=f'"{cat_options}"', allow_blank=True)
    dv_cat.error = "Choose a category"
    ws_tx.add_data_validation(dv_cat)
    dv_cat.add(f"{cat_letter}2:{cat_letter}{tx_data_end}")

    # Subcategory dropdown: dependent on category (INDIRECT uses same-row category)
    sub_formula = f'=INDIRECT(SUBSTITUTE(SUBSTITUTE(SUBSTITUTE(INDIRECT(ADDRESS(ROW(),{cat_col}))," & ","_")," ","_"),"-","_"))'
    dv_sub = DataValidation(type="list", formula1=sub_formula, allow_blank=True)
    dv_sub.error = "Choose a subcategory for the selected category"
    ws_tx.add_data_validation(dv_sub)
    dv_sub.add(f"{sub_letter}2:{sub_letter}{tx_data_end}")

    # --- Summary sheet (formulas linked to Transactions) ---
    ws_sum = wb.create_sheet("Summary", 3)
    ws_sum["A1"] = "Summary"
    ws_sum["A1"].font = Font(bold=True)
    ws_sum["A2"] = "Total transactions"
    ws_sum["B2"] = n_tx
    ws_sum["A3"] = "Total inflow"
    ws_sum["B3"] = f"=SUMIF(Transactions!{amount_letter}2:{amount_letter}{tx_data_end},\">0\")"
    ws_sum["A4"] = "Total outflow"
    ws_sum["B4"] = f"=ABS(SUMIF(Transactions!{amount_letter}2:{amount_letter}{tx_data_end},\"<0\"))"
    ws_sum["A5"] = "Net"
    ws_sum["B5"] = f"=SUM(Transactions!{amount_letter}2:{amount_letter}{tx_data_end})"

    # By category and subcategory: all combinations from config + (choose), with SUMIFS formulas
    ws_sum["A7"] = "By category and subcategory"
    ws_sum["A7"].font = Font(bold=True)
    ws_sum["A8"] = "Category"
    ws_sum["B8"] = "Subcategory"
    ws_sum["C8"] = "Amount"
    ws_sum["A8"].font = Font(bold=True)
    ws_sum["B8"].font = Font(bold=True)
    ws_sum["C8"].font = Font(bold=True)

    sum_rows = [("(choose)", "(choose)")]  # matches empty category/subcategory in Transactions
    for cat, subs in categories.items():
        for sub in subs:
            sum_rows.append((cat, sub))
    for idx, (cat, sub) in enumerate(sum_rows):
        r = 9 + idx
        ws_sum[f"A{r}"] = cat
        ws_sum[f"B{r}"] = sub
        # (choose) matches blank cells; others match exact category/subcategory
        cat_crit = '""' if cat == "(choose)" else f'A{r}'
        sub_crit = '""' if sub == "(choose)" else f'B{r}'
        ws_sum[f"C{r}"] = f'=SUMIFS(Transactions!${amount_letter}$2:${amount_letter}${tx_data_end},Transactions!${cat_letter}$2:${cat_letter}${tx_data_end},{cat_crit},Transactions!${sub_letter}$2:${sub_letter}${tx_data_end},{sub_crit})'

    # --- Pivot sheet: placeholder (real pivot added by xlwings below) ---
    ws_pivot = wb.create_sheet("Pivot", 4)
    ws_pivot["A1"] = "Pivot Table (refreshes when Excel opens)"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(EXCEL_OUTPUT)

    # Add real Excel pivot table + pivot chart (requires Excel installed)
    _add_excel_pivot_table(str(EXCEL_OUTPUT.absolute()), n_tx)


def main():
    csv_path = find_first_csv()
    if not csv_path:
        print(f"No CSV file found in {INPUT_DIR}/")
        print("Please add a Belfius CSV export to the input folder.")
        return 1

    if not CONFIG_FILE.exists():
        print(f"Categories config not found: {CONFIG_FILE}")
        return 1

    with open(CONFIG_FILE, "r", encoding="utf-8") as f:
        categories = json.load(f)

    print(f"Reading {csv_path}...")
    raw = parse_belfius_csv(csv_path)
    df = clean_and_normalize(raw)

    # Add category columns
    df["category"] = ""
    df["subcategory"] = ""
    for i in range(len(df)):
        cat, sub = apply_rules(df.iloc[i], categories)
        df.at[df.index[i], "category"] = cat
        df.at[df.index[i], "subcategory"] = sub

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Save clean CSV
    df.to_csv(CSV_OUTPUT, index=False, sep=";", encoding="utf-8")
    print(f"Saved {CSV_OUTPUT}")

    # Save Excel dashboard
    create_excel_dashboard(df, categories)
    print(f"Saved {EXCEL_OUTPUT}")

    print("Done.")
    return 0


if __name__ == "__main__":
    exit(main())
